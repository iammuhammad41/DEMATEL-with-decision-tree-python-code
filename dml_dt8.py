import numpy as np
import matplotlib.pyplot as plt
import os
import glob
import csv
from openpyxl import Workbook
import pandas as pd
from sklearn.tree import DecisionTreeClassifier, plot_tree
from sklearn.linear_model import LogisticRegression
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.decomposition import PCA
from sklearn.model_selection import KFold, cross_val_score, StratifiedKFold
from sklearn.metrics import classification_report, confusion_matrix, accuracy_score
from sklearn.preprocessing import StandardScaler
from sklearn.feature_selection import RFE
from imblearn.over_sampling import SMOTE
import shap
import warnings
from matplotlib.lines import Line2D

# Suppress warnings for cleaner output
warnings.filterwarnings('ignore')


class DEMATELSolver:
    def __init__(self):
        self.experts = []
        self.factors = []
        self.numberOfExperts = 0
        self.numberOfFactors = 0
        self.matrix = []
        self.result = {"cause": [], "effect": []}
        self.consistency_metrics = {}
        self.Z = None
        self.X = None
        self.T = None
        self.R = None
        self.C = None
        self.prominence = None
        self.relation = None
        self.feature_importances = None
        self.models = {}

    # Setters and Getters
    def setNumberOfExperts(self, nb):
        self.numberOfExperts = nb

    def getNumberOfExperts(self):
        return self.numberOfExperts

    def setNumberOfFactors(self, nb):
        self.numberOfFactors = nb

    def getNumberOfFactors(self):
        return self.numberOfFactors

    def getMatrix(self):
        return self.matrix

    def setMatrix(self, matrices):
        self.matrix = matrices

    def AddMatrix(self, matrix):
        self.matrix.append(matrix)

    def setExperts(self, experts):
        self.experts = experts

    def getExperts(self):
        return self.experts

    def addExpert(self, expert):
        self.experts.append(expert)

    def setFactors(self, factors):
        self.factors = factors

    def getFactors(self):
        return self.factors

    def addFactor(self, factor):
        self.factors.append(factor)

    def getDirectInfluenceMatrix(self):
        return self.Z

    def getNormalizedDirectInfluenceMatrix(self):
        return self.X

    def getTotalInfluenceMatrix(self):
        return self.T

    def getProminence(self):
        return self.R + self.C

    def getRelation(self):
        return self.R - self.C

    # Methods for reading files
    def read_matrix_from_csv(self, file_path):
        """
        Reads a single direct influence matrix from a CSV file.
        Assumes that the CSV contains only numerical values.
        """
        try:
            with open(file_path, 'r', newline='') as csvfile:
                reader = csv.reader(csvfile)
                matrix = []
                for row in reader:
                    # Convert each row to a list of floats
                    matrix.append([float(value.strip()) for value in row if value.strip() != ''])
                return np.array(matrix)
        except Exception as e:
            print(f"Error reading {file_path}: {e}")
            return None

    def read_multiple_matrices_from_csv_folder(self, folder_path):
        """
        Reads multiple direct influence matrices from CSV files in a folder.
        Each CSV file should represent one expert's matrix.
        Assumes all matrices are of the same size.
        """
        csv_files = glob.glob(os.path.join(folder_path, "*.csv"))
        if not csv_files:
            print("No CSV files found in the provided folder.")
            return

        for file in csv_files:
            matrix = self.read_matrix_from_csv(file)
            if matrix is not None:
                self.matrix.append(matrix)
                expert_name = os.path.splitext(os.path.basename(file))[0]
                self.experts.append(expert_name)
                print(f"Loaded matrix for expert: {expert_name}")

        self.numberOfExperts = len(self.experts)
        if len(self.matrix) > 0:
            self.numberOfFactors = self.matrix[0].shape[0]
            print(f"Number of experts loaded: {self.numberOfExperts}")
            print(f"Number of factors loaded: {self.numberOfFactors}")
        else:
            print("No matrices were loaded.")

    def read_factors_from_file(self, file_path):
        """
        Reads factor names from a text file, one factor per line.
        """
        try:
            with open(file_path, 'r') as f:
                factors = [line.strip() for line in f if line.strip() != '']
            self.factors = factors
            self.numberOfFactors = len(factors)
            print(f"Number of factors loaded: {self.getNumberOfFactors()}")
        except Exception as e:
            print(f"Error reading factors from {file_path}: {e}")

    # DEMATEL Steps
    def step1(self, expert_weights=None):
        """
        Step 1: Aggregate the direct influence matrices from all experts to form matrix Z.
        Z = (Sum of all matrices * weights) / Number of experts
        If expert_weights is provided, use them; otherwise, assume equal weighting.
        """
        if self.numberOfExperts == 0:
            print("No experts' matrices loaded. Cannot perform Step 1.")
            return
        self.Z = np.zeros((self.numberOfFactors, self.numberOfFactors))
        for idx, matrix in enumerate(self.matrix):
            expert = self.experts[idx]
            weight = expert_weights.get(expert,
                                        1 / self.numberOfExperts) if expert_weights else 1 / self.numberOfExperts
            self.Z += weight * matrix
        print("Step 1 Completed: Aggregated Direct Influence Matrix (Z)")

    def step2(self):
        """
        Step 2: Normalize the direct influence matrix Z to obtain matrix X.
        X = Z / S, where S is the maximum row or column sum of Z.
        """
        if self.Z is None:
            print("Matrix Z is not computed. Cannot perform Step 2.")
            return
        self.X = self.Z / self.calculateS()
        print("Step 2 Completed: Normalized Direct Influence Matrix (X)")

    def step3(self):
        """
        Step 3: Calculate the total influence matrix T.
        T = X * (I - X)^-1
        """
        if self.X is None:
            print("Matrix X is not computed. Cannot perform Step 3.")
            return
        I = np.eye(self.numberOfFactors)
        try:
            self.T = np.dot(self.X, np.linalg.inv(I - self.X))
            print("Step 3 Completed: Total Influence Matrix (T)")
        except np.linalg.LinAlgError:
            print("Matrix (I - X) is singular and cannot be inverted. Step 3 failed.")
            self.T = None

    def step4(self):
        """
        Step 4: Determine cause and effect factors based on R and C.
        R = sum of rows of T
        C = sum of columns of T
        Prominence = R + C
        Relation = R - C
        If Relation > 0: Cause Factor
        Else: Effect Factor
        """
        if self.T is None:
            print("Total Influence Matrix T is not computed. Skipping Step 4.")
            return

        self.R = np.sum(self.T, axis=1)
        self.C = np.sum(self.T, axis=0)
        self.prominence = self.R + self.C
        self.relation = self.R - self.C

        # Assign Cause and Effect factors based on Relation
        for i, relation_value in enumerate(self.relation):
            if relation_value > 0:
                self.result["cause"].append(self.factors[i])
            else:
                self.result["effect"].append(self.factors[i])

        print("Step 4 Completed: Identified Cause and Effect Factors")
        print(f"Cause Factors: {self.result['cause']}")
        print(f"Effect Factors: {self.result['effect']}")

    def calculateS(self):
        """
        Calculates the normalization factor S.
        S = max(max row sum of Z, max column sum of Z)
        """
        row_sums = np.sum(self.Z, axis=1)
        col_sums = np.sum(self.Z, axis=0)
        S = max(np.max(row_sums), np.max(col_sums))
        print(f"Normalization Factor (S): {S}")
        return S

    def drawCurve(self):
        """
        Draws the DEMATEL Influential Relation Map (IRM).
        """
        if self.prominence is None or self.relation is None:
            print("Prominence or Relation not calculated. Cannot draw curve.")
            return

        plt.figure(figsize=(12, 8))

        # Determine colors based on relation
        colors = ['red' if r < 0 else 'green' for r in self.relation]

        plt.scatter(self.prominence, self.relation, marker='o', s=100, c=colors, alpha=0.6)

        for name, p, r in zip(self.factors, self.prominence, self.relation):
            plt.text(p, r, name, ha='center', va='bottom', fontsize=9)

        plt.xlabel('Prominence (R + C)')
        plt.ylabel('Relation (R - C)')
        plt.title('DEMATEL Influential Relation Map (IRM)')
        plt.axhline(0, color='black', linestyle='--')
        plt.axvline(0, color='black', linestyle='--')
        plt.grid(True)

        # Create custom legend
        legend_elements = [
            Line2D([0], [0], marker='o', color='w', label='Cause Factor', markerfacecolor='green', markersize=10),
            Line2D([0], [0], marker='o', color='w', label='Effect Factor', markerfacecolor='red', markersize=10)
        ]
        plt.legend(handles=legend_elements, loc='upper right')

        plt.tight_layout()
        plt.show()

    def savexl(self, url):
        """
        Saves the DEMATEL analysis results to an Excel file.
        """
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Influential Relation Map IRM"
        ws1.append(["Factor name", "R", "C", "R+C", "R-C"])

        for nameFactor, r, c, rpc, rmc in zip(self.factors, self.R, self.C, self.prominence, self.relation):
            ws1.append([nameFactor, r, c, rpc, rmc])

        # Function to write a matrix to a worksheet
        def write_matrix(ws, title, matrix):
            ws.title = title
            ws.append([""] + self.factors)  # Header
            for i, row in enumerate(matrix, start=1):
                ws.append([self.factors[i - 1]] + list(row))

        # Direct Influence Matrix
        ws2 = wb.create_sheet(title="Direct-influence matrix")
        write_matrix(ws2, "Direct-influence matrix", self.Z)

        # Normalized Direct Influence Matrix
        ws3 = wb.create_sheet(title="Normalized Influence matrix")
        write_matrix(ws3, "Normalized Influence matrix", self.X)

        # Total Influence Matrix
        if self.T is not None:
            ws4 = wb.create_sheet(title="Total-influence matrix")
            write_matrix(ws4, "Total-influence matrix", self.T)

        # Save the workbook
        try:
            os.makedirs(url, exist_ok=True)  # Create directory if it doesn't exist
            file_path = os.path.join(url, "DEMATELAnalysis.xlsx")
            wb.save(file_path)
            print(f"Excel file successfully created at {file_path}.")
        except Exception as e:
            print(f"Error saving Excel file: {e}")

    def calculate_consistency_metrics(self):
        """
        Calculates consistency metrics based on inter-expert variance.
        """
        if self.numberOfExperts < 2:
            print("Insufficient number of experts to calculate consistency metrics.")
            return

        # Stack all expert matrices into a 3D array: (number_of_experts, factors, factors)
        stacked_matrices = np.array(self.matrix)  # Shape: (E, F, F)

        # Calculate variance across experts for each element
        variance_matrix = np.var(stacked_matrices, axis=0)

        # Calculate mean variance as overall consistency measure
        mean_variance = np.mean(variance_matrix)

        # Calculate standard deviation
        std_variance = np.std(variance_matrix)

        print(f"Mean Variance Across Experts: {mean_variance:.4f}")
        print(f"Standard Deviation of Variance: {std_variance:.4f}")

        # Store for future use
        self.consistency_metrics = {
            "variance_matrix": variance_matrix,
            "mean_variance": mean_variance,
            "std_variance": std_variance
        }

    # Enhanced Feature Selection
    def select_features_rfe(self, X, y, n_features_to_select=6):
        """
        Selects top features using Recursive Feature Elimination (RFE) with Logistic Regression.
        """
        model = LogisticRegression(solver='liblinear')
        rfe = RFE(model, n_features_to_select=n_features_to_select)
        fit = rfe.fit(X, y)
        selected_features = X.columns[fit.support_].tolist()
        print(f"Selected Features by RFE: {selected_features}")
        return selected_features

    def select_features_random_forest(self, X, y, n_features_to_select=6):
        """
        Selects top features using feature importances from Random Forest.
        """
        model = RandomForestClassifier(random_state=42)
        model.fit(X, y)
        importances = pd.Series(model.feature_importances_, index=X.columns)
        selected_features = importances.sort_values(ascending=False).head(n_features_to_select).index.tolist()
        print(f"Selected Features by Random Forest: {selected_features}")
        return selected_features

    # Enhanced Model Training with Multiple Models and k-Fold CV
    def build_models(self, X, y):
        """
        Builds and evaluates multiple machine learning models using k-Fold Cross-Validation.
        Models included: Decision Tree, Logistic Regression, SVM, k-NN, Naive Bayes, Random Forest, Gradient Boosting
        """
        models = {
            'Decision Tree': DecisionTreeClassifier(random_state=42, max_depth=5, min_samples_leaf=2),
            'Logistic Regression': LogisticRegression(solver='liblinear', penalty='l2', C=1.0),
            'SVM': SVC(kernel='linear', probability=True, random_state=42),
            'k-NN': KNeighborsClassifier(n_neighbors=3),
            'Naive Bayes': GaussianNB(),
            'Random Forest': RandomForestClassifier(random_state=42, n_estimators=100),
            'Gradient Boosting': GradientBoostingClassifier(random_state=42, n_estimators=100)
        }

        # Handle class imbalance with SMOTE
        smote = SMOTE(random_state=42)
        X_resampled, y_resampled = smote.fit_resample(X, y)

        # Feature Scaling
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X_resampled)

        # Initialize k-Fold Cross-Validation
        kf = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)

        # Dictionary to store results
        results = {}

        for name, model in models.items():
            print(f"\nBuilding and evaluating model: {name}")
            cv_scores = cross_val_score(model, X_scaled, y_resampled, cv=kf, scoring='accuracy')
            results[name] = cv_scores
            print(f"Average Accuracy: {cv_scores.mean():.4f} (+/- {cv_scores.std():.4f})")

            # Fit the model on the entire resampled dataset
            model.fit(X_scaled, y_resampled)
            self.models[name] = model

            # Feature Importance or Coefficients
            if hasattr(model, 'feature_importances_'):
                importances = pd.Series(model.feature_importances_, index=X.columns)
                print(f"Feature Importances:\n{importances.sort_values(ascending=False)}")
            elif hasattr(model, 'coef_'):
                coefficients = pd.Series(model.coef_[0], index=X.columns)
                print(f"Model Coefficients:\n{coefficients.sort_values(ascending=False)}")
            else:
                print("No feature importances or coefficients available for this model.")

        return results

    def interpret_models_with_shap(self, X, y):
        """
        Uses SHAP to interpret feature importances for various models.
        """
        # Feature Scaling
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)

        for name, model in self.models.items():
            if isinstance(model, (RandomForestClassifier, GradientBoostingClassifier, DecisionTreeClassifier)):
                print(f"\nInterpreting model: {name} with SHAP")
                explainer = shap.TreeExplainer(model)
                shap_values = explainer.shap_values(X_scaled)

                # Plot SHAP summary
                shap.summary_plot(shap_values, X, plot_type="bar")
            elif isinstance(model, LogisticRegression):
                print(f"\nInterpreting model: {name} with SHAP")
                # Corrected parameter
                explainer = shap.LinearExplainer(model, X_scaled, feature_perturbation="interventional")
                shap_values = explainer.shap_values(X_scaled)
                shap.summary_plot(shap_values, X, plot_type="bar")
            elif isinstance(model, SVC):
                print(f"\nInterpreting model: {name} with SHAP")
                # Note: SVC with probability=True is computationally expensive with SHAP
                explainer = shap.KernelExplainer(model.predict_proba, X_scaled)
                shap_values = explainer.shap_values(X_scaled[:100])  # Limit for speed; adjust as needed
                shap.summary_plot(shap_values, X[:100], plot_type="bar")
            else:
                print(f"SHAP not implemented for model: {name}")


def main():
    # Hardcoded paths
    factors_path = "./inputs/factors.txt"
    matrices_path = "./inputs/matrices"
    output_path = "./outputs"
    data_path = "./inputs/expert_data.csv"
    target_variable_column = "Cluster"  # Ensure this column exists in expert_data.csv

    # Initialize DEMATELSolver
    solver = DEMATELSolver()

    # Read factor names
    print(f"Reading factors from {factors_path}...")
    solver.read_factors_from_file(factors_path)

    # Read experts' matrices
    print(f"Reading experts' matrices from {matrices_path}...")
    solver.read_multiple_matrices_from_csv_folder(matrices_path)

    # Check if matrices are loaded
    if not solver.matrix:
        print("No matrices loaded. Exiting.")
        return

    # Calculate Consistency Metrics
    print("\nCalculating Consistency Metrics...")
    solver.calculate_consistency_metrics()

    # Execute DEMATEL steps
    print("\nExecuting DEMATEL steps...")
    solver.step1()
    print("\nDirect Influence Matrix (Z):\n", solver.Z)

    solver.step2()
    print("\nNormalized Direct Influence Matrix (X):\n", solver.X)

    solver.step3()
    if solver.T is not None:
        print("\nTotal Influence Matrix (T):\n", solver.T)
    else:
        print("\nTotal Influence Matrix (T) could not be computed due to singular matrix.")

    solver.step4()
    if solver.T is not None:
        print("\nRelation (R - C):\n", solver.relation)
        print("\nProminence (R + C):\n", solver.prominence)
        print("\nCause Factors:", solver.result["cause"])
        print("Effect Factors:", solver.result["effect"])

    # Draw the DEMATEL curve
    print("\nDrawing DEMATEL Influential Relation Map...")
    solver.drawCurve()

    # Save to Excel
    print("\nSaving results to Excel...")
    solver.savexl(output_path)

    # Check Class Balance
    cause_count = len(solver.result["cause"])
    effect_count = len(solver.result["effect"])
    print(f"\nClass Distribution - Cause: {cause_count}, Effect: {effect_count}")

    # Prepare Feature Selection
    print("\nRefining Feature Selection using RFE and Random Forest...")
    # Assuming that you have a dataset prepared with all factors as samples and selected features as columns
    # Here, we'll create a temporary dataset for feature selection purposes

    # Initialize a dictionary to hold influence scores for all factors
    factors = solver.getFactors()
    influence_scores = {factor: [] for factor in factors}

    # Read each expert matrix and aggregate influence scores for all features
    csv_files = glob.glob(os.path.join(matrices_path, "*.csv"))
    for file in csv_files:
        try:
            matrix = np.loadtxt(file, delimiter=',')
            for factor in factors:
                factor_idx = factors.index(factor)
                # Using all features initially for feature selection
                influence = matrix[factor_idx, :]
                influence_scores[factor].append(influence)
        except Exception as e:
            print(f"Error reading {file}: {e}")
            continue

    # Calculate average influence scores for all features
    avg_influence_scores = {factor: np.mean(scores, axis=0) for factor, scores in influence_scores.items()}

    # Create DataFrame with all features
    data = []
    for factor in factors:
        try:
            row = avg_influence_scores[factor]
            # Assign cluster based on whether the factor is a Cause or Effect
            cluster = 'Cause' if factor in solver.result["cause"] else 'Effect'
            data.append(list(row) + [cluster])
        except KeyError:
            print(f"Influence scores for factor '{factor}' not found.")
            continue

    columns = factors + ['Cluster']
    df_all = pd.DataFrame(data, columns=columns)

    # Handle any missing data
    df_all.dropna(inplace=True)

    # Encode target variable
    df_all['Cluster'] = df_all['Cluster'].map({'Cause': 1, 'Effect': 0})

    # Feature Selection using RFE
    X_all = df_all[factors]
    y_all = df_all['Cluster']
    selected_features_rfe = solver.select_features_rfe(X_all, y_all, n_features_to_select=6)

    # Feature Selection using Random Forest
    selected_features_rf = solver.select_features_random_forest(X_all, y_all, n_features_to_select=6)

    # Combine selected features from both methods
    selected_features = list(set(selected_features_rfe + selected_features_rf))
    print(f"Combined Selected Features: {selected_features}")

    # Update the dataset with selected features
    data = []
    for factor in factors:
        if factor in selected_features:
            try:
                # Since 'influence_scores[factor]' is a list of arrays, we take the mean across all experts
                row = avg_influence_scores[factor]
                # Assign cluster based on whether the factor is a Cause or Effect
                cluster = 'Cause' if factor in solver.result["cause"] else 'Effect'
                data.append(list(row) + [cluster])
            except IndexError:
                print(f"Influence score for factor '{factor}' not found.")
                continue
        else:
            # For simplicity, we're selecting only the selected features as columns
            # In practice, you should include all selected features' influence scores
            pass

    # Create DataFrame with selected features
    # Here, for demonstration, we'll assume selected_features are distinct and not overlapping
    # Adjust accordingly based on actual selected features
    df_selected = df_all[selected_features + ['Cluster']]

    # Save to CSV
    df_selected.to_csv(data_path, index=False)
    print(f"\nSelected features dataset created at {data_path}.")

    # Check Class Balance again after feature selection
    cause_count = df_selected['Cluster'].sum()
    effect_count = len(df_selected) - cause_count
    print(f"\nClass Distribution after Feature Selection - Cause: {int(cause_count)}, Effect: {int(effect_count)}")

    # Build and Evaluate Models with Enhanced Techniques
    print("\nBuilding and evaluating machine learning models with enhanced techniques...")
    X = df_selected[selected_features]
    y = df_selected['Cluster']

    # Initialize DEMATELSolver models
    results = solver.build_models(X, y)

    # Interpret Models with SHAP
    print("\nInterpreting models with SHAP for interpretability...")
    solver.interpret_models_with_shap(X, y)

    # Continuous Validation Placeholder
    print("\nSetting up Continuous Validation...")
    print("Ensure to periodically retrain and validate models with new data to maintain performance.")


if __name__ == "__main__":
    main()
